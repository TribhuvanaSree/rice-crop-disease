import openpyxl
import random
 
# Define variable to load the dataframe

dataframe = openpyxl.load_workbook(r"D:\epics work\BookMois.xlsx")
# Define variable to read sheet
dataframe1 = dataframe.active

r=random.randrange(50)
x=0
for col in dataframe1.iter_cols(1, dataframe1.max_column):
    x+=1
    flag=1
    if col[r].value<25 :
        flag = 0
        break
if flag:
    print(col[r].value)
else:
    print("Not ok")
