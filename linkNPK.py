import openpyxl
import random
 
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(r"D:\epics work\Book1.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active
l=[]
# Iterate the loop to read the cell valuesw32
z = {1:0.3,2:2,3:1.6}
r=random.randrange(50)
x=0
for col in dataframe1.iter_cols(1, dataframe1.max_column):
    l.append(col[r].value)
    x+=1
    flag=1
    if col[r].value<z[x] :
        flag = 0
        break
if flag:
    print("N: ",l[0],"\nP: ",l[1],"\nk: ",l[2])
else:
    print("Not ok",col[r].value)
